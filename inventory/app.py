"""
Flask Inventory Management Application

The application is a web base business inventory management system built with Flask

Features include:
    User authentication with hashed passwords
    Inventory management with use of database queries (insert, delete, update)
    Exporting of user items via XML and/or XLSX files
    Low/No stock tracking
    Image uploading
"""


from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import sqlite3
import os
import base64
import xml.etree.ElementTree as ET
import io
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, PatternFill, Font
import tempfile

app = Flask(__name__, template_folder=os.path.join(os.path.dirname(__file__), "..", "templates"))
app.secret_key = os.urandom(12)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


# Changed to support both databases
def init_database():
    """
    Initialize the database using sql script from 'inventory_schema.sql'

    Function creates user and inventory tables if they have not been created
    """
    with sqlite3.connect(f'inventory.db') as connection:
        with open(f'inventory_schema.sql') as f:
            connection.executescript(f.read())



class User(UserMixin):
    """
    User class for Flask-Login

    Attributes:
        id (int): User id from database
        username (str): Username of account
        user_password (str): User password hashed from database
        store_name (str): Store name of account
    """
    def __init__(self, user_id, username, user_password, store_name):
        self.id = user_id
        self.username = username
        self.user_password = user_password
        self.store_name = store_name

    # Function returns a User class with the values stored in the user database
    @staticmethod
    def get(user_id):
        """
        Retrieve user from the database using ID

        Args:
            user_id (int): ID of user

        Returns:
            User | None: An instance of User class if it exists in the database, otherwise None
        """
        connection = sqlite3.connect('inventory.db')
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        user = cursor.execute('SELECT * FROM users WHERE user_id = ?', (user_id,)).fetchone()
        if user is None:
            return None
        else:
            return User(user['user_id'], user['username'], user['user_password'], user['store_name'])


@login_manager.user_loader
def load_user(user_id):
    """
    Flask-login user loader

    Args:
        user_id (str): Takes the ID of a user as a string

    Returns:
        User | None: An instance of User class if it exists in the database, otherwise None
    """
    return User.get(int(user_id))


@app.route('/home')
def home():
    """
    Display the home page after login.

    Returns:
        Response: Rendered home template.
        """
    return render_template('home.html')


# Converts a file to binary
def convert_to_binary(filename):
    """
    Converts a file via it's name to binary

    Args:
         filename (str): Name of the file path to be converted to binary

    Returns:
        bytes: Binary of the file
    """
    with open(filename, 'rb') as f:
        blob_data = f.read()
    return blob_data


# Takes the database table and downloads a xml file to the users computer
@app.route('/xml-export')
@login_required
def inventory_to_xml():
    """
    Export the current inventory to an XML file.

    Returns:
        Response: XML file download
    """
    connection = sqlite3.connect('inventory.db')
    cursor = connection.cursor()
    cursor.execute('SELECT name, image, description, quantity, price FROM inventory WHERE owner_id = ?',
                   (current_user.id,))
    # Stores all the rows of the database
    rows = cursor.fetchall()
    connection.close()

    '''
    Creates a tree and creates a child node (item) for each item and then adds a child node (characteristics [name,
    image, quantity, etc.]) for each column in a row. The value of the characteristic is stored in the
    corresponding node
    '''
    root = ET.Element('inventory')
    for row in rows:
        item_elmt = ET.SubElement(root, 'item')
        name_elmt = ET.SubElement(item_elmt, 'name')
        name_elmt.text = row[0]
        image_elmt = ET.SubElement(item_elmt, 'image')
        '''
        Have to check if an image is stored, for this item.
        If an image is stored then have to convert from binary back into base64 
        '''
        if row[1]:
            image_base64 = base64.b64encode(row[1]).decode('utf-8')
            image_elmt.text = image_base64
        else:
            image_elmt.text = ''
        description_elmt = ET.SubElement(item_elmt, 'description')
        description_elmt.text = row[2]
        quantity_elmt = ET.SubElement(item_elmt, 'quantity')
        quantity_elmt.text = str(row[3])
        price_elmt = ET.SubElement(item_elmt, 'price')
        price_elmt.text = f'{row[4]:.2f}'

    # Sets up output file
    output_xml = io.BytesIO()
    tree = ET.ElementTree(root)
    tree.write(output_xml, encoding='utf-8', xml_declaration=True)
    output_xml.seek(0)

    return send_file(output_xml, mimetype='application/xml', as_attachment=True, download_name='inventory.xml')


@app.route('/xlsx-export')
@login_required
def inventory_to_xlsx():
    """
    Export the current inventory to an XLSX file

    Returns:
         Response: XLSX file download
    """
    connection = sqlite3.connect('inventory.db')
    cursor = connection.cursor()
    cursor.execute('SELECT name, image, description, quantity, price FROM inventory WHERE owner_id = ?',
                   (current_user.id,))
    rows = cursor.fetchall()
    connection.close()

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Write headers
    headers = ["Name", "Image", "Description", "Quantity", "Price"]
    ws.append(headers)

    # Style headers: bold, white text, blue fill, centered
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    row_index = 2
    for name, image_blob, description, quantity, price in rows:
        ws.cell(row=row_index, column=1, value=name)
        ws.cell(row=row_index, column=3, value=description)
        ws.cell(row=row_index, column=4, value=quantity)
        ws.cell(row=row_index, column=5, value=f'{price:.2f}')

        # Center all text fields
        for col_num in [1, 3, 4, 5]:
            ws.cell(row=row_index, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_index].height = 60  # for images

        if image_blob:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            temp_file.write(image_blob)
            temp_file.close()

            img = XLImage(temp_file.name)
            img.height = 60
            img.width = 60
            ws.add_image(img, f"B{row_index}")

        row_index += 1

    output_xlsx = io.BytesIO()
    wb.save(output_xlsx)
    output_xlsx.seek(0)

    return send_file(
        output_xlsx,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='inventory.xlsx'
    )


@app.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    """
    Add an item to the inventory of the current user

    GET:
        Render the add item form
    POST:
        Insert item into inventory table

    Returns:
        str | Response: Render template or redict to inventory page
    """
    if request.method == 'POST':
        name = request.form['name']
        image_file = request.files.get('image')
        description = request.form['description']
        quantity = int(request.form['quantity'])
        price = float(request.form['price'])
        #
        image_blob = None

        if image_file and image_file.filename != '':
            # Temporarily save the file to read it as binary
            temp_folder = 'temp'
            os.makedirs(temp_folder, exist_ok=True)

            temp_path = os.path.join(temp_folder, image_file.filename)
            image_file.save(temp_path)
            image_blob = convert_to_binary(temp_path)
            os.remove(temp_path)
        try:
            with sqlite3.connect('inventory.db') as items:
                cursor = items.cursor()
                cursor.execute('INSERT INTO INVENTORY (name, image, description, quantity, price, owner_id) \
                            VALUES (?, ?, ?, ?, ?, ?)',
                               (name, image_blob, description, quantity, price, current_user.id))
                items.commit()
            return redirect(url_for('home'))
        except sqlite3.IntegrityError:
            flash("Item already found in inventory. Please change the name.");
            return render_template("add.html");
    else:
        return render_template('add.html')


@app.route('/inventory')
@login_required
def inventory():
    """
    Display the inventory of the current user

    Returns:
        str: Render HTML template with current user's items
    """
    connection = sqlite3.connect('inventory.db')
    cursor = connection.cursor()
    cursor.execute('SELECT name, image, description, quantity, price FROM inventory WHERE owner_id = ?',
                   (current_user.id,))
    rows = cursor.fetchall()
    connection.close()

    data = []
    for row in rows:
        name, image_blob, description, quantity, price = row
        if image_blob:
            image_base64 = base64.b64encode(image_blob).decode('utf-8')
            image_uri = f"data:image/jpeg;base64,{image_base64}"
        else:
            image_uri = None  # or a default placeholder image path

        data.append((name, image_uri, description, quantity, price))

    return render_template('inventory.html', data=data)


# New main page. Checks user database to see if login information exists.
# If login information exists moves to the home page and creates a user class
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Handle user login

    GET:
         Renders the login page

    POST:
        Checks inputted user info and logs them in if it exists

    Returns:
        str | Response: Render template or redict to home page
    """
    if request.method == 'POST':
        username = request.form['username']
        user_password = request.form['user_password']
        connection = sqlite3.connect('inventory.db')
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()
        user_row = cursor.execute('SELECT *  FROM USERS WHERE username = ?', (username,)).fetchone()
        connection.close()
        if user_row and check_password_hash(user_row['user_password'], user_password):
            user = User(user_row['user_id'], user_row['username'], user_row['user_password'], user_row['store_name'])
            login_user(user)
            return redirect(url_for('home'))
        else:
            return render_template('login.html')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """
    Logout the current user

    Returns:
        Response: Redirect to login page
    """
    logout_user()
    return redirect(url_for('login'))


# Users are able to register new accounts. Usernames have to be unique and passwords are stored as a hashed password.
# New users are stored in the database and are able to log in.
@app.route('/register', methods=['GET', 'POST'])
def register():
    """
    Handle user registration

    GET:
        Render the registration form
    POST:
        Insert a new user into the database with hashed password

    Returns:
        str | Response: Rendered template or redirect response
    """
    if request.method == 'POST':
        username = request.form['username'].strip()
        raw_password = request.form['user_password']
        store_name = request.form['store_name'].strip()

        # Always hash once, right before storing
        hashed_password = generate_password_hash(raw_password)

        # Use a context manager to ensure commit/close happen cleanly
        with sqlite3.connect('inventory.db') as connection:
            connection.row_factory = sqlite3.Row
            cursor = connection.cursor()

            # Check for existing user by USERNAME ONLY (unique constraint handles true duplicates)
            existing = cursor.execute(
                'SELECT 1 FROM USERS WHERE username = ?',
                (username,)
            ).fetchone()
            if existing:
                flash('Username already taken. Please choose another.')
                return render_template('register.html')

            cursor.execute(
                'INSERT INTO USERS (username, user_password, store_name) VALUES (?, ?, ?)',
                (username, hashed_password, store_name)
            )
            # context manager will commit automatically unless an exception occurs

        # Option 1: send them to login page immediately
        # return redirect(url_for('login'))

        # Option 2 (recommended UX): log them in right away after registration
        with sqlite3.connect('inventory.db') as connection:
            connection.row_factory = sqlite3.Row
            cursor = connection.cursor()
            row = cursor.execute(
                'SELECT * FROM USERS WHERE username = ?',
                (username,)
            ).fetchone()

        #
        return redirect(url_for('login'))

    return render_template('register.html')


# Adam created this function
# If the users tries to update the value a sql query is made updating the value in the database.
# If a get request is sent the edit_quantity html is rendered and the item name and quantity stored in the database
# is displayed.
@app.route('/edit/<string:name>', methods=['GET', 'POST'])
def edit_quantity(name):
    """
    Edit an existing inventory item.

    Args:
        name (str): name of the item to edit

    GET:
        Render the edit form with existing item data
    POST:
        Update the item in the database

    Returns:
        str | Response: Rendered template or redirect to inventory page
    """
    conn = sqlite3.connect('inventory.db')
    cursor = conn.cursor()

    if request.method == 'POST':
        new_quantity = int(request.form['quantity'])
        cursor.execute('UPDATE inventory SET quantity = ? WHERE name = ?', (new_quantity, name))
        conn.commit()
        conn.close()
        return redirect(url_for('inventory'))

    # GET request
    cursor.execute('SELECT quantity FROM inventory WHERE name = ?', (name,))
    row = cursor.fetchone()
    conn.close()

    if row:
        current_quantity = row[0]
        return render_template('edit_quantity.html', name=name, quantity=current_quantity)
    else:
        return "Item not found", 404


@app.route('/low-stock')
@login_required
def low_stock():
    """
    Displays all items with quantity less than 11

    Returns:
        str: Render low stock template
    """
    connection = sqlite3.connect('inventory.db')
    cursor = connection.cursor()
    cursor.execute('SELECT name, image, description, quantity, price FROM inventory WHERE owner_id = ?',
                   (current_user.id,))
    rows = cursor.fetchall()
    connection.close()

    lowStock = []
    outOfStock = []
    for row in rows:
        name, image_blob, description, quantity, price = row
        if image_blob:
            image_base64 = base64.b64encode(image_blob).decode('utf-8')
            image_uri = f"data:image/jpeg;base64,{image_base64}"
        else:
            image_uri = None  # or a default placeholder image path

        if row[3] == 0:
            outOfStock.append((name, image_uri, description, quantity, price))
        elif row[3] <= 10:
            lowStock.append((name, image_uri, description, quantity, price))

    return render_template('low_stock.html', lowStock=lowStock, outOfStock=outOfStock)


@app.route('/delete', methods=['GET', 'POST'])
@login_required
def delete():
    """
    Delete an inventory item

    GET:
        Render the delete form

    POST:
        Delete the specified item

    Returns:
        str | Response: Rendered template or redirect to inventory page
    """
    connection = sqlite3.connect('inventory.db')
    cursor = connection.cursor()
    cursor.execute('SELECT name, item_id FROM inventory WHERE owner_id = ?',
                   (current_user.id,))
    rows = cursor.fetchall()

    if request.method == 'POST':
        itemToDelete = int(request.form['deleteItem'])
        cursor.execute('DELETE from inventory WHERE item_id = ?', (itemToDelete,))
        connection.commit()
        connection.close()
        return redirect(url_for('inventory'))

    items = []
    for row in rows:
        name, item_id = row
        items.append((name, item_id))

    connection.close()
    return render_template('delete.html', items=items)


if __name__ == '__main__':
    init_database()
    app.run(debug=True)

